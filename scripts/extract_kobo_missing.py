from __future__ import annotations

import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


WORKBOOK_GLOB = "TerremotoVE*_all_versions*.xlsx"
ROOT = Path(__file__).resolve().parents[1]
DESKTOP = Path.home() / "Desktop"
OUTPUT_DIR = ROOT / "data"
SOURCE_LABEL = "KoBoToolbox 2026-06-26"

PHONE_RE = re.compile(r"(?:\+?\d[\d\-\s\(\)]{6,}\d)")
EMAIL_RE = re.compile(r"[\w.\-+]+@[\w.\-]+\.\w+")
HANDLE_RE = re.compile(r"@\w{3,}")
URL_RE = re.compile(r"https?://\S+")
CI_RE = re.compile(r"(?:C\.?I\.?|cedula|c[eé]dula)\s*[:\-]?\s*([\d.]{5,})", re.IGNORECASE)
AGE_RE = re.compile(r"(\d{1,3})\s*a[nñ]os?", re.IGNORECASE)

NAME_TOKEN = r"[A-ZÁÉÍÓÚÑ][a-záéíóúñ]+|[A-ZÁÉÍÓÚÑ]{2,}"
NAME_RE = re.compile(rf"\b({NAME_TOKEN}(?:\s+{NAME_TOKEN}){{1,4}})\b")

GENERIC_NON_PERSON = (
    "edificio colapsado",
    "edificio moises derrumbado",
    "lista legible",
    "personas atendidas",
    "hospital perez carre",
    "de 2 a 4 desaparecidos",
    "dos personas desaparecidas",
    "hotel la mar suites",
    "aeropuerto",
    "22 heridos",
    "36 heridos",
    "locales afectados",
)

ANIMAL_KEYWORDS = (
    "gata",
    "gato",
    "perro",
    "mascota",
    "animales",
)

STOP_NAMES = {
    "La Guaira",
    "Catia La Mar",
    "Playa Grande",
    "Los Corales",
    "Tanaguarenas",
    "Caraballeda",
    "Maiquetia",
    "Maiquetía",
    "Macuto",
    "Tucacas",
    "Universidad Maritima Del Caribe",
    "Universidad Marítima Del Caribe",
    "Residencias Caribe",
    "Edificio El Caribe",
    "Costa Mar",
    "Costa Brava",
    "Punta Piedras",
}

LOW_CONFIDENCE_NAME_PARTS = {
    "auromar",
    "caraballeda",
    "tahiti",
    "guaira",
    "pareja",
    "ucv",
    "palace",
    "opp",
    "res",
    "edificio",
    "residencias",
}


@dataclass
class Candidate:
    firstName: str
    lastName: str
    documentId: str
    age: str
    gender: str
    lastSeen: str
    description: str
    photoUrl: str
    contactName: str
    contactPhone: str
    sourceId: int


def normalize(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def clean_ci(value: str) -> str:
    return re.sub(r"\D", "", value)


def looks_like_url(value: str) -> bool:
    return value.startswith("http://") or value.startswith("https://")


def split_name(full_name: str) -> tuple[str, str] | None:
    tokens = [token for token in full_name.strip().split() if token]
    if len(tokens) < 2:
        return None
    return tokens[0].title(), " ".join(tokens[1:]).title()


def find_first_contact(texts: Iterable[str]) -> str:
    for text in texts:
        scrubbed = URL_RE.sub(" ", text)
        for phone_match in PHONE_RE.finditer(scrubbed):
            digits = re.sub(r"\D", "", phone_match.group(0))
            if 7 <= len(digits) <= 15:
                return normalize(phone_match.group(0))

    for text in texts:
        email_match = EMAIL_RE.search(text)
        if email_match:
            return email_match.group(0)

    for text in texts:
        handle_match = HANDLE_RE.search(text)
        if handle_match:
            return handle_match.group(0)

    for text in texts:
        url_match = URL_RE.search(text)
        if url_match:
            return url_match.group(0)

    return ""


def find_contact_name(text: str) -> str:
    lowered = text.lower()
    patterns = [
        r"contacto de la familia[:\s]+([A-ZÁÉÍÓÚÑ][^.,;]+)",
        r"otro contacto[^:]*[:\s]+([A-ZÁÉÍÓÚÑ][^.,;]+)",
        r"contactar a su [^ ]+\s+([A-ZÁÉÍÓÚÑ][^(\n,.;]+)",
        r"contacto[:\s]+([A-ZÁÉÍÓÚÑ][^0-9,.;]+)",
    ]
    if "contact" not in lowered and "llamar" not in lowered and "comunicar" not in lowered:
        return "Contacto del reporte"

    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            candidate = normalize(match.group(1).strip(" :-"))
            parsed = split_name(candidate)
            if parsed:
                return candidate.title()

    return "Contacto del reporte"


def infer_last_seen(description: str, extra: str, link: str, location: str) -> str:
    combined = " ".join(part for part in [description, extra, link] if part)
    patterns = [
        r"[Vv]isto por ultima vez en ([^.,;]+)",
        r"[ÚUu]ltima ubicaci[oó]n conocida(?: fue)? ([^.,;]+)",
        r"[Dd]esaparecido en ([^.,;]+)",
        r"[Ee]staba en ([^.,;]+)",
        r"[Rr]eside en ([^.,;]+)",
        r"[Vv]ive en ([^.,;]+)",
        r"[Ee]ncontraba en ([^.,;]+)",
        r"[Ee]n ([A-ZÁÉÍÓÚÑa-záéíóúñ0-9\s\-]+, [A-ZÁÉÍÓÚÑa-záéíóúñ\s\-]+)",
    ]

    for pattern in patterns:
        match = re.search(pattern, combined)
        if match:
            return normalize(match.group(1)).strip(".")

    if link and not looks_like_url(link):
        return link

    if location:
        return f"Ubicacion reportada: {location}"

    return "Ubicacion no especificada"


def extract_ci(text: str) -> str:
    match = CI_RE.search(text)
    if match:
        return clean_ci(match.group(1))
    return ""


def extract_age(text: str) -> str:
    match = AGE_RE.search(text)
    if match:
        return match.group(1)
    return ""


def cleanup_for_names(text: str) -> str:
    cleaned = URL_RE.sub(" ", text)
    cleaned = PHONE_RE.sub(" ", cleaned)
    cleaned = EMAIL_RE.sub(" ", cleaned)
    cleaned = cleaned.replace("C.I.", " ").replace("CI", " ").replace("Cédula", " ")
    cleaned = re.sub(r"\([^)]*\)", " ", cleaned)
    cleaned = cleaned.replace("—", " ").replace("/", " ")
    return normalize(cleaned)


def normalize_name(value: str) -> str:
    return " ".join(part.title() for part in normalize(value).split())


def extract_leading_name(segment: str) -> list[str]:
    stripped = normalize(segment)
    prefixes = [
        r"^desaparecid[oa]s?:?\s*",
        r"^desaparecido el señor\s*",
        r"^desaparecida la señora\s*",
        r"^desaparecida:\s*",
        r"^desaparecido niño,\s*",
        r"^persona desaparecida:?\s*",
        r"^familia de \w+ desaparecida[^:]*:\s*",
        r"^familia [\w\s]+ desaparecida[^:]*:\s*",
    ]
    for prefix in prefixes:
        stripped = re.sub(prefix, "", stripped, flags=re.IGNORECASE)

    stop_words = [
        " se encuentra",
        " vive en",
        " estaba en",
        " visto por ultima vez",
        " reside en",
        " en ",
        " - ",
        ",",
        ".",
    ]
    end = len(stripped)
    lowered = stripped.lower()
    for marker in stop_words:
        position = lowered.find(marker)
        if position > 0:
            end = min(end, position)
    candidate = stripped[:end].strip(" .,-:")
    if split_name(candidate):
        return [normalize_name(candidate)]
    return []


def extract_names(description: str, extra: str) -> list[str]:
    found: list[str] = []
    combined = " ".join(part for part in [description, extra] if part)

    found.extend(extract_leading_name(description))

    cleaned = cleanup_for_names(combined)
    for match in NAME_RE.finditer(cleaned):
        name = normalize_name(match.group(1))
        if name in STOP_NAMES:
            continue
        if split_name(name) is None:
            continue
        if all(part.lower() in {"la", "guaira", "catia", "mar", "playa", "grande", "macuto", "maiquetia", "caraballeda"} for part in name.split()):
            continue
        found.append(name)

    unique: list[str] = []
    seen = set()
    for item in found:
        key = item.lower()
        if key not in seen:
            seen.add(key)
            unique.append(item)
    return unique


def should_skip(description: str, extra: str) -> bool:
    lowered = f"{description} {extra}".lower()
    if any(keyword in lowered for keyword in ANIMAL_KEYWORDS):
        return True
    if any(keyword in lowered for keyword in GENERIC_NON_PERSON):
        return True
    return False


def build_description(raw_description: str, extra: str, link: str, source_id: int) -> str:
    parts = [raw_description]
    if extra:
        parts.append(f"Informacion adicional: {extra}")
    if link:
        parts.append(f"Referencia: {link}")
    parts.append(f"Fuente: {SOURCE_LABEL} · ID {source_id}")
    return " | ".join(normalize(part) for part in parts if normalize(part))


def find_workbook() -> Path:
    matches = sorted(DESKTOP.glob(WORKBOOK_GLOB))
    if not matches:
        raise FileNotFoundError(f"No se encontro un archivo que coincida con {WORKBOOK_GLOB}")
    return matches[-1]


def main() -> None:
    workbook_path = find_workbook()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    rows = list(sheet.iter_rows(values_only=True))
    headers = list(rows[0])
    idx = {header: position for position, header in enumerate(headers)}

    person_key = headers[5]
    family_key = headers[12]
    description_key = headers[13]
    link_key = headers[14]
    location_key = headers[15]
    extra_key = headers[20]
    id_key = headers[21]

    imported: list[dict[str, str | int]] = []
    safe_imported: list[dict[str, str | int]] = []
    review: list[dict[str, str | int]] = []
    seen_records = set()

    for row in rows[1:]:
        if row[idx[person_key]] != 1 and row[idx[family_key]] != 1:
            continue

        description = normalize(row[idx[description_key]])
        link = normalize(row[idx[link_key]])
        location = normalize(row[idx[location_key]])
        extra = normalize(row[idx[extra_key]])
        source_id = int(row[idx[id_key]])

        if not description:
            review.append(
                {
                    "sourceId": source_id,
                    "reason": "Sin descripcion",
                    "description": description,
                    "extra": extra,
                    "link": link,
                }
            )
            continue

        if should_skip(description, extra):
            review.append(
                {
                    "sourceId": source_id,
                    "reason": "No parece un caso importable de persona desaparecida",
                    "description": description,
                    "extra": extra,
                    "link": link,
                }
            )
            continue

        names = extract_names(description, extra)
        if not names:
            review.append(
                {
                    "sourceId": source_id,
                    "reason": "No se pudo extraer un nombre confiable",
                    "description": description,
                    "extra": extra,
                    "link": link,
                }
            )
            continue

        contact_phone = find_first_contact([extra, link, description])
        contact_name = find_contact_name(f"{extra} {description}")
        last_seen = infer_last_seen(description, extra, link, location)
        document_id = extract_ci(description) or extract_ci(extra)
        age = extract_age(description) or extract_age(extra)

        if not contact_phone:
            review.append(
                {
                    "sourceId": source_id,
                    "reason": "Sin contacto usable para el formulario",
                    "description": description,
                    "extra": extra,
                    "link": link,
                    "names": names,
                }
            )
            continue

        for full_name in names:
            split = split_name(full_name)
            if not split:
                continue

            first_name, last_name = split
            dedupe_key = (
                first_name.lower(),
                last_name.lower(),
                clean_ci(document_id),
            )
            if dedupe_key in seen_records:
                continue
            seen_records.add(dedupe_key)

            record = Candidate(
                firstName=first_name,
                lastName=last_name,
                documentId=clean_ci(document_id),
                age=age,
                gender="Sin especificar",
                lastSeen=last_seen,
                description=build_description(description, extra, link, source_id),
                photoUrl="",
                contactName=contact_name,
                contactPhone=contact_phone,
                sourceId=source_id,
            )
            serialized = record.__dict__
            imported.append(serialized)

            lowered_name_parts = {
                part.lower()
                for part in f"{record.firstName} {record.lastName}".split()
            }
            has_bad_name = bool(lowered_name_parts & LOW_CONFIDENCE_NAME_PARTS)
            has_url_contact = looks_like_url(record.contactPhone)
            has_numeric_tail = bool(re.search(r"\d", record.lastName))
            if not has_bad_name and not has_url_contact and not has_numeric_tail:
                safe_imported.append(serialized)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    imported_path = OUTPUT_DIR / "kobo-missing-people.json"
    safe_imported_path = OUTPUT_DIR / "kobo-missing-people-safe.json"
    review_path = OUTPUT_DIR / "kobo-missing-people-review.json"
    summary_path = OUTPUT_DIR / "kobo-missing-people-summary.json"

    imported_path.write_text(json.dumps(imported, ensure_ascii=False, indent=2), encoding="utf-8")
    safe_imported_path.write_text(json.dumps(safe_imported, ensure_ascii=False, indent=2), encoding="utf-8")
    review_path.write_text(json.dumps(review, ensure_ascii=False, indent=2), encoding="utf-8")
    summary_path.write_text(
        json.dumps(
            {
                "sourceWorkbook": str(workbook_path),
                "sourceRowsMarkedMissing": sum(
                    1
                    for row in rows[1:]
                    if row[idx[person_key]] == 1 or row[idx[family_key]] == 1
                ),
                "formReadyRecords": len(imported),
                "safeFormReadyRecords": len(safe_imported),
                "manualReviewRows": len(review),
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    print(
        json.dumps(
            {
                "sourceWorkbook": str(workbook_path),
                "formReadyRecords": len(imported),
                "safeFormReadyRecords": len(safe_imported),
                "manualReviewRows": len(review),
                "importedPath": str(imported_path),
                "safeImportedPath": str(safe_imported_path),
                "reviewPath": str(review_path),
                "summaryPath": str(summary_path),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
